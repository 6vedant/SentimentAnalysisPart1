import pandas as pd
import griddb_python as griddb
import openpyxl

list_sheetnames = pd.ExcelFile("sample_dataset.xlsx").sheet_names
total_num_sheets = len(list_sheetnames)

for i in range(total_num_sheets):
    sheet_read = pd.read_excel("sample_dataset.xlsx", list_sheetnames[i])
    tweet_data_df = pd.DataFrame(sheet_read)

    factory = griddb.StoreFactory.get_instance()

    # Get GridStore object
    # Provide the necessary arguments
    gridstore = factory.get_store(
        host=argv[1],
        port=int(argv[2]),
        cluster_name=argv[3],
        username=argv[4],
        password=argv[5]
    )

    # Define the container names
    tweet_dataaset_container = list_sheetnames[i]

    # Get the containers
    tweet_data = gridstore.get_container(tweet_dataaset_container)

    # Fetch all rows - circuits_container
    query = tweet_data.query("select *")
    rs = query.fetch(False)
    print(f"{tweet_dataaset_container} Data")

    # Iterate and create a list
    retrieved_data = []
    while rs.has_next():
        data = rs.next()
        retrieved_data.append(data)

    print(retrieved_data)

    # Convert the list to a pandas data frame
    tweet_dataframe = pd.DataFrame(retrieved_data,
                                      columns=['sno', 'twitter_name', 'twitter_id', 'tweet', 'date'])

    # Get the data frame details
    print(tweet_dataframe)
    tweet_dataframe.info()



