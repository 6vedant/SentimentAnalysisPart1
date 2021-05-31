import pandas as pd
import openpyxl
import griddb_python as griddb
import sys


list_sheetnames = pd.ExcelFile("sample_dataset.xlsx").sheet_names
total_num_sheets = len(list_sheetnames)

for i in range(total_num_sheets):
    sheet_read = pd.read_excel("sample_dataset.xlsx", list_sheetnames[i])
    df = pd.DataFrame(sheet_read)
    tweets_list = sheet_read["Tweet"].values.tolist()
    tweet_date_list = sheet_read["Date"].values.tolist()
    tweet_author_names_list = sheet_read["Screen Name"].values.tolist()
    tweet_author_handle_list = sheet_read["Twitter Id"].values.tolist()
    print(tweet_author_handle_list[:5])
    print(tweet_author_names_list[:5])

# a function to clean the tweets using regular expression
def clean_tweet(tweet):
    '''
    Utility function to clean the text in a tweet by removing
    links and special characters using regex.
    '''
    return ' '.join(re.sub("(@[A-Za-z0-9]+)|([^0-9A-Za-z \t])|(\w+:\/\/\S+)", " ", tweet).split())


circuits_container = "circuits"
races_container = "races"

# Put rows
    # Define the data frames
    circuits_columns.put_rows(circuits_data)

for i in range(total_num_sheets):
    sheet_read = pd.read_excel("sample_dataset.xlsx", list_sheetnames[i])
    tweet_data_df = pd.DataFrame(sheet_read)
    tweets_list = sheet_read["Tweet"].values.tolist()
    tweet_date_list = sheet_read["Date"].values.tolist()
    tweet_author_names_list = sheet_read["Screen Name"].values.tolist()
    tweet_author_handle_list = sheet_read["Twitter Id"].values.tolist()
    curr_sheet = list_sheetnames[i]

    factory = griddb.StoreFactory.get_instance()

    # Get GridStore object
    # Provide the necessary arguments
    gridstore = factory.get_store(
        host=argv[1],
        port=int(argv[2]),
        cluster_name=argv[3],
        username=argv[4],
        password=argv[5]
    )

    curr_sheet_griddb_container = curr_sheet

    # create collection for the tweet data in the sheet
    tweet_data_container_info = griddb.ContainerInfo(circuits_container,
                                                     [["sno", griddb.Type.INTEGER],
                                                      ["twitter_name", griddb.Type.STRING],
                                                      ["twitter_id", griddb.Type.STRING],
                                                      ["tweet", griddb.Type.STRING],
                                                      ["date", griddb.Type.STRING]],
                                                     griddb.ContainerType.COLLECTION, True)

    tweets_columns = gridstore.put_container(tweet_data_container_info)
    # Put rows
    # Pass the data frames as param
    tweets_columns.put_rows(tweet_data_df)
